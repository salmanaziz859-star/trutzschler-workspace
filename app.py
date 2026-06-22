import os
import json

import urllib.request
import webbrowser
# =====================================================================
# SALMAN'S AUTO-UPDATER ENGINE (START)
# =====================================================================
import sys
import urllib.request

# 1. Aapka GitHub Repository ka Direct Raw Link
GITHUB_RAW_URL = "https://raw.githubusercontent.com/YOUR_GITHUB_USERNAME/YOUR_REPO_NAME/main/app.py"
LOCAL_FILE = __file__  # Yeh is file ka apna local path khud pakad lega

def check_for_updates():
    print("Checking for Salman's new features...")
    try:
        temp_file = LOCAL_FILE + ".tmp"
        # Background mein chupke se naya code download karna
        urllib.request.urlretrieve(GITHUB_RAW_URL, temp_file)
        
        # Agar sahi download ho jaye toh purani file ko replace kar dena
        if os.path.getsize(temp_file) > 0:
            if os.path.exists(LOCAL_FILE):
                os.remove(LOCAL_FILE)
            os.rename(temp_file, LOCAL_FILE)
            print("Software updated successfully with new functions!")
    except Exception as e:
        print("Running on last saved features (Offline Mode)...")

# Software chalte hi sabse pehle update check karega
check_for_updates()
# =====================================================================
# SALMAN'S AUTO-UPDATER ENGINE (END)
# =====================================================================
from datetime import datetime
import customtkinter as ctk
from tkinter import messagebox, filedialog, ttk, simpledialog
from PIL import Image, ImageTk, ImageEnhance

# ReportLab core imports fix for NumberedCanvas
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
from reportlab.lib.styles import ParagraphStyle

# openpyxl imports for Excel engine
from openpyxl import Workbook
from openpyxl.styles import Font as XLFont, PatternFill, Alignment, Border, Side

ctk.set_appearance_mode("System")
ctk.set_default_color_theme("blue")

ADMIN_USER = "admin"
ADMIN_PASS = "1234"

# --- PERMANENT STORAGE SETUP ---
APP_DATA_DIR = os.path.join(os.environ.get('LOCALAPPDATA', os.path.expanduser('~')), "TrutzschlerWorkspace")
BASE_HISTORY_DIR = os.path.join(APP_DATA_DIR, "System_History")
DB_FILE_PATH = os.path.join(APP_DATA_DIR, "inventory_database.json")

for directory in [APP_DATA_DIR, BASE_HISTORY_DIR]:
    if not os.path.exists(directory):
        os.makedirs(directory)

BG_IMAGE_PATH = os.path.join(APP_DATA_DIR, "trutzschler_bg.png")

# --- Automatic Background Image Downloader ---
if not os.path.exists(BG_IMAGE_PATH):
    try:
        IMAGE_URL = "https://images.unsplash.com/photo-1581091226825-a6a2a5aee158?q=80&w=1280&auto=format&fit=crop"
        headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64)'}
        req = urllib.request.Request(IMAGE_URL, headers=headers)
        with urllib.request.urlopen(req) as response, open(BG_IMAGE_PATH, 'wb') as out_file:
            out_file.write(response.read())
    except Exception as e:
        print(f"Network asset sync skipped: {e}")

# --- DEFAULT INVENTORY DATA WITH PERMANENT BASE PRICES ---
DEFAULT_INVENTORY = {
    "Blowroom": [
        {"id": "1.1.1.1", "name": "Condenser", "code": "BR-COI", "specs": "Standard condenser for material transport and air separation", "price": "12500"},
        {"id": "1.1.1.2", "name": "Stock Trunk", "code": "FD-S 1200", "specs": "High capacity stock trunk for material buffering", "price": "18400"},
        {"id": "1.1.1.3", "name": "2 Way Distribution", "code": "BR-2W", "specs": "Pneumatic distribution system with active change boxes", "price": "9200"},
        {"id": "1.1.1.4", "name": "Automatic Change Box", "code": "BR-AC", "specs": "High performance control change box for multi-line distribution", "price": "14500"}
    ],
    "Card": [
        {"id": "1.2.1.1", "name": "Card Clothing Wire", "code": "CR-WR", "specs": "High durability cylinder and doffer wire set", "price": "4200"},
        {"id": "1.2.1.2", "name": "Flat Tops", "code": "CR-FT", "specs": "Precision engineered flat tops for enhanced trash removal", "price": "6800"}
    ],
    "Drawing": [
        {"id": "1.3.1.1", "name": "Top Roller", "code": "DR-TR", "specs": "Precision top roller with heavy duty bearing alignments", "price": "1500"},
        {"id": "1.3.1.2", "name": "Bottom Fluted Roller", "code": "DR-BR", "specs": "High tolerance bottom fluted mechanical roller", "price": "2900"}
    ],
    "Comber": [
        {"id": "1.4.1.1", "name": "Top Comb", "code": "CM-TC", "specs": "Premium pinning density top comb segment", "price": "2100"},
        {"id": "1.4.1.2", "name": "Circular Comb", "code": "CM-CC", "specs": "Graded combing profile circular comb block assembly", "price": "5400"}
    ]
}

def load_inventory_from_db():
    if os.path.exists(DB_FILE_PATH):
        try:
            with open(DB_FILE_PATH, "r") as f:
                data = json.load(f)
                for cat in data:
                    for item in data[cat]:
                        if "price" not in item:
                            item["price"] = ""
                return data
        except Exception:
            return DEFAULT_INVENTORY
    else:
        with open(DB_FILE_PATH, "w") as f:
            json.dump(DEFAULT_INVENTORY, f, indent=4)
        return DEFAULT_INVENTORY

def save_inventory_to_db(data):
    with open(DB_FILE_PATH, "w") as f:
        json.dump(data, f, indent=4)


class NumberedCanvas(canvas.Canvas):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self._saved_page_states = []
    def showPage(self):
        self._saved_page_states.append(dict(self.__dict__))
        self._startPage()
    def save(self):
        num_pages = len(self._saved_page_states)
        for state in self._saved_page_states:
            self.__dict__.update(state)
            self.draw_decorations(num_pages)
            super().showPage()
        super().save()
    def draw_decorations(self, total_pages):
        self.saveState()
        
        self.setFont("Helvetica-Bold", 14)
        self.setFillColor(colors.HexColor("#003399"))
        self.drawString(54, 755, "TRÜTZSCHLER")
        
        self.setFont("Helvetica-Bold", 14)
        self.setFillColor(colors.HexColor("#7f8c8d"))
        self.drawString(162, 755, "SPINNING")
        
        self.setFont("Helvetica", 7.5)
        self.setFillColor(colors.HexColor("#555555"))
        self.drawString(54, 743, "Trützschler Group SE  PO Box 410164  D-41241 Mönchengladbach")
        
        self.setStrokeColor(colors.HexColor("#003399"))
        self.setLineWidth(0.75)
        self.line(54, 732, 558, 732)
        
        self.setFont("Helvetica", 6.5)
        self.setFillColor(colors.HexColor("#333333"))
        self.setStrokeColor(colors.HexColor("#cccccc"))
        self.setLineWidth(0.5)
        self.line(54, 75, 558, 75)
        
        col1_y = 65
        self.drawString(54, col1_y, "Board of Directors:")
        self.drawString(54, col1_y - 9, "Heinrich Krull")
        self.drawString(54, col1_y - 18, "Florian Rück")
        self.drawString(54, col1_y - 27, "Alexander Stampfer")
        
        col2_x = 160
        self.drawString(col2_x, col1_y, "Chairman of the Supervisory Board:")
        self.drawString(col2_x, col1_y - 9, "Dr.-Ing. Roland Münch")
        self.drawString(col2_x, col1_y - 18, "Trützschler Group SE, AG Mönchengladbach HRB 20308")
        self.drawString(col2_x, col1_y - 27, "Headquarters: Mönchengladbach Deutschland")
        
        col3_x = 345
        self.drawString(col3_x, col1_y, "Deutsche Bank AG, Mönchengladbach")
        self.drawString(col3_x, col1_y - 9, "IBAN: DE 85 3107 0001 0720 9000 00")
        self.drawString(col3_x, col1_y - 18, "BIC: DEUTDEDD310")
        self.drawString(col3_x, col1_y - 27, "VAT ID No.: DE 291 704 386")
        
        col4_x = 465
        self.drawString(col4_x, col1_y, "Commerzbank AG")
        self.drawString(col4_x, col1_y - 9, "Mönchengladbach")
        self.drawString(col4_x, col1_y - 18, "IBAN: DE 66 3104 0015 0363 6008 00")
        self.drawString(col4_x, col1_y - 27, "BIC: COBADEFF310")
        
        self.setFont("Helvetica", 8)
        self.drawRightString(558, 30, f"Page {self._pageNumber} of {total_pages}")
        self.restoreState()


class LoginWindow(ctk.CTkFrame):
    def __init__(self, master, on_success):
        super().__init__(master, fg_color="#0b0f19")
        self.on_success = on_success
        self.pack(expand=True, fill="both")
        
        self.bg_label = ctk.CTkLabel(self, text="")
        self.bg_label.place(x=0, y=0, relwidth=1, relheight=1)
        self.bind("<Configure>", self.resize_background)
        
        self.box = ctk.CTkFrame(self, width=420, height=480, corner_radius=24, fg_color="#121826", border_width=1.5, border_color="#2a364f")
        self.box.place(relx=0.5, rely=0.5, anchor="center")
        self.box.pack_propagate(False) 
        
        ctk.CTkLabel(self.box, text="TRÜTZSCHLER", font=("Segoe UI", 32, "bold"), text_color="#3b82f6").pack(pady=(45, 2))
        ctk.CTkLabel(self.box, text="GERMAN ENGINEERING WORKSPACE", font=("Segoe UI", 10, "bold"), text_color="#64748b").pack(pady=(0, 35))
        
        self.ent_user = ctk.CTkEntry(self.box, placeholder_text="Username", width=320, height=44, font=("Segoe UI", 14), corner_radius=10, fg_color="#090d16", border_color="#1e293b", text_color="#f8fafc", placeholder_text_color="#475569")
        self.ent_user.pack(pady=12)
        
        self.ent_pass = ctk.CTkEntry(self.box, placeholder_text="Password", show="*", width=320, height=44, font=("Segoe UI", 14), corner_radius=10, fg_color="#090d16", border_color="#1e293b", text_color="#f8fafc", placeholder_text_color="#475569")
        self.ent_pass.pack(pady=12)
        
        self.ent_pass.bind("<Return>", lambda event: self.check_credentials())
        self.ent_user.bind("<Return>", lambda event: self.check_credentials())

        self.btn_access = ctk.CTkButton(self.box, text="SECURE ACCESS SYSTEM", font=("Segoe UI", 13, "bold"), width=320, height=46, corner_radius=10, fg_color="#3b82f6", hover_color="#2563eb", text_color="#ffffff")
        self.btn_access.configure(command=self.check_credentials)
        self.btn_access.pack(pady=(35, 20))

    def resize_background(self, event=None):
        if os.path.exists(BG_IMAGE_PATH):
            try:
                window_width = self.winfo_width()
                window_height = self.winfo_height()
                if window_width > 10 and window_height > 10:
                    img = Image.open(BG_IMAGE_PATH)
                    img = img.resize((window_width, window_height), Image.Resampling.LANCZOS)
                    enhancer = ImageEnhance.Brightness(img)
                    img = enhancer.enhance(0.45) 
                    self.bg_img_ctk = ctk.CTkImage(light_image=img, dark_image=img, size=(window_width, window_height))
                    self.bg_label.configure(image=self.bg_img_ctk)
            except Exception as e:
                print(f"Sizing hook feedback: {e}")
        else:
            self.bg_label.configure(fg_color="#0b0f19")

    def check_credentials(self):
        if self.ent_user.get().strip() == ADMIN_USER and self.ent_pass.get().strip() == ADMIN_PASS:
            self.pack_forget()
            self.on_success()
        else:
            messagebox.showerror("Access Denied", "Invalid Credentials. Please try again.")
            
    def clear_fields(self):
        self.ent_user.delete(0, "end")
        self.ent_pass.delete(0, "end")
        self.ent_user.focus()


class CustomInputDialog(ctk.CTkToplevel):
    def __init__(self, parent, title, prompts, initial_values=None):
        super().__init__(parent)
        self.title(title)
        self.geometry("400x420") 
        self.resizable(False, False)
        
        # Windows system configurations for stable focus mapping
        self.transient(parent)
        self.lift()
        
        self.result = None
        self.entries = {}
        
        ctk.CTkLabel(self, text=title, font=("Arial", 16, "bold"), text_color="#1f538d").pack(pady=15)
        
        first_ent = None
        for placeholder, key in prompts:
            frame = ctk.CTkFrame(self, fg_color="transparent")
            frame.pack(fill="x", padx=30, pady=5)
            
            # FIXED: Explicit normal state to guarantee layout usability
            ent = ctk.CTkEntry(frame, placeholder_text=placeholder, width=340, height=35, state="normal")
            ent.pack()
            
            if initial_values and key in initial_values:
                ent.insert(0, initial_values[key])
                
            self.entries[key] = ent
            if first_ent is None:
                first_ent = ent
            
        btn_frame = ctk.CTkFrame(self, fg_color="transparent")
        btn_frame.pack(fill="x", padx=30, pady=20)
        
        ctk.CTkButton(btn_frame, text="Cancel", width=100, fg_color="#64748b", command=self.destroy).pack(side="left", padx=5)
        ctk.CTkButton(btn_frame, text="Save Item", width=100, fg_color="#228B22", command=self.validate_and_save).pack(side="right", padx=5)
        
        self.center_window()
        
        # FIXED: Removed the buggy "<Button-1>" dialog focus-stealing binding entirely.
        # Enforcing immediate structural focus on window generation
        self.focus_force()
        self.grab_set()
        if first_ent:
            first_ent.after(200, lambda: first_ent.focus_set())

    def center_window(self):
        self.update_idletasks()
        width = self.winfo_width()
        height = self.winfo_height()
        x = (self.winfo_screenwidth() // 2) - (width // 2)
        y = (self.winfo_screenheight() // 2) - (height // 2)
        self.geometry(f'{width}x{height}+{x}+{y}')

    def validate_and_save(self):
        res = {k: v.get().strip() for k, v in self.entries.items()}
        if not res["name"]:
            messagebox.showwarning("Fields Required", "Item Name is a mandatory field.")
            return
        if not res["specs"]:
            res["specs"] = "No additional technical specification provided."
        self.result = res
        self.destroy()


class MasterSystemApp(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("Trutzschler Proforma & Quotation Matrix Workspace")
        self.geometry("1320x850")
        
        self.inventory_db = load_inventory_from_db()
        
        self.sidebar = None
        self.main_workspace = None
        self.grid_inputs = {}
        self.sub_scroll_frames = {}
        self.app_loaded = False  
        
        # Safe binding global click handler
        self.bind("<Button-1>", self.remove_global_cursor_focus)
        
        self.login_view = LoginWindow(self, self.load_main_application)

    def remove_global_cursor_focus(self, event):
        if not self.app_loaded:
            return
            
        try:
            # FIXED: Do not alter global workspace structural focus if a top-level popup is active
            for child in self.winfo_children():
                if isinstance(child, ctk.CTkToplevel) and child.winfo_exists():
                    return

            widget_under_mouse = self.winfo_containing(event.x_root, event.y_root)
            widget_str = str(widget_under_mouse).lower()
            if "entry" in widget_str or "text" in widget_str:
                return  
                
            self.focus_set()
        except Exception:
            pass

    def load_main_application(self):
        self.app_loaded = True 
        self.grid_columnconfigure(0, weight=0, minsize=290)
        self.grid_columnconfigure(1, weight=1)
        self.grid_rowconfigure(0, weight=1)

        self.sidebar = ctk.CTkFrame(self, width=290, corner_radius=0)
        self.sidebar.grid(row=0, column=0, sticky="nsew")
        
        self.main_workspace = ctk.CTkFrame(self, fg_color="transparent")
        self.main_workspace.grid(row=0, column=1, sticky="nsew", padx=20, pady=20)
        
        self.sidebar_setup()
        self.workspace_setup()
        self.toggle_dynamic_fields()

    def sidebar_setup(self):
        ctk.CTkLabel(self.sidebar, text="TRÜTZSCHLER", font=("Arial", 22, "bold"), text_color="#1f538d").pack(pady=(20, 2))
        ctk.CTkLabel(self.sidebar, text="Enterprise Tracking Core", font=("Arial", 11, "italic"), text_color="gray").pack(pady=(0, 15))
        ctk.CTkLabel(self.sidebar, text="📋 DOCUMENT CONFIGURATION", font=("Arial", 11, "bold")).pack(anchor="w", padx=20, pady=(10, 2))
        
        self.doc_type_var = ctk.StringVar(value="Quotation")
        combo_type = ctk.CTkComboBox(self.sidebar, values=["Quotation", "Proforma"], variable=self.doc_type_var, width=250, font=("Arial", 12, "bold"), command=self.toggle_dynamic_fields)
        combo_type.pack(padx=20, pady=5)

        self.ent_factory = ctk.CTkEntry(self.sidebar, placeholder_text="Client Factory Name (AGI Denim)", width=250)
        self.ent_factory.pack(padx=20, pady=4)
        self.ent_address = ctk.CTkEntry(self.sidebar, placeholder_text="Full Factory Address Location", width=250)
        self.ent_address.pack(padx=20, pady=4)
        self.ent_offer = ctk.CTkEntry(self.sidebar, placeholder_text="Offer / Ref (e.g., 02-013106-00)", width=250)
        self.ent_offer.pack(padx=20, pady=4)
        self.ent_project = ctk.CTkEntry(self.sidebar, placeholder_text="Project Name (Draw Frames)", width=250)
        self.ent_project.pack(padx=20, pady=4)

        self.lbl_previews = ctk.CTkLabel(self.sidebar, text="⚙️ WORKING ENGINE PREVIEWS", font=("Arial", 11, "bold"))
        self.lbl_previews.pack(anchor="w", padx=20, pady=(15, 2))
        
        self.btn_preview_pdf = ctk.CTkButton(self.sidebar, text="🔍 Preview Quotation (PDF)", font=("Arial", 13, "bold"), fg_color="#ffcc00", text_color="black", hover_color="#e6b800", height=38, command=self.trigger_pdf_preview)
        self.btn_preview_excel = ctk.CTkButton(self.sidebar, text="🔍 Preview Spreadsheet (Excel)", font=("Arial", 13, "bold"), fg_color="#ffcc00", text_color="black", hover_color="#e6b800", height=38, command=self.trigger_excel_preview)

        self.lbl_exports = ctk.CTkLabel(self.sidebar, text="📦 EXPORT & STATE TRACKING", font=("Arial", 11, "bold"))
        self.lbl_exports.pack(anchor="w", padx=20, pady=(20, 2))
        
        self.btn_export_excel = ctk.CTkButton(self.sidebar, text="📊 Save & Export Quotation Excel", font=("Arial", 13, "bold"), fg_color="#1f538d", height=38, command=self.generate_excel_format)
        self.btn_export_pdf = ctk.CTkButton(self.sidebar, text="📜 Save & Export Quotation PDF", font=("Arial", 13, "bold"), fg_color="#228B22", hover_color="#006400", height=38, command=self.generate_pdf_proforma)

        ctk.CTkLabel(self.sidebar, text="🔒 SECURITY SESSIONS", font=("Arial", 11, "bold")).pack(anchor="w", padx=20, pady=(25, 2))
        btn_logout = ctk.CTkButton(self.sidebar, text="🔴 Logout System", font=("Arial", 13, "bold"), fg_color="#cc3333", hover_color="#992222", height=35, command=self.trigger_system_logout)
        btn_logout.pack(padx=20, pady=4, fill="x")

    def toggle_dynamic_fields(self, *args):
        doc_type = self.doc_type_var.get()
        if doc_type == "Quotation":
            self.ent_address.pack_forget()
            self.ent_project.pack_forget()
            self.ent_offer.pack(padx=20, pady=4, after=self.ent_factory)
            self.btn_preview_pdf.pack_forget()
            self.btn_export_pdf.pack_forget()
            self.btn_preview_excel.pack(padx=20, pady=4, fill="x", after=self.lbl_previews)
            self.btn_export_excel.pack(padx=20, pady=4, fill="x", after=self.lbl_exports)
            self.btn_preview_excel.configure(text="🔍 Preview Spreadsheet (Excel)")
            self.btn_export_excel.configure(text="📊 Save & Export Quotation Excel")
        else:
            self.ent_address.pack_forget()
            self.ent_offer.pack_forget()
            self.ent_project.pack_forget()
            self.ent_address.pack(padx=20, pady=4, after=self.ent_factory)
            self.ent_offer.pack(padx=20, pady=4, after=self.ent_address)
            self.ent_project.pack(padx=20, pady=4, after=self.ent_offer)
            self.btn_preview_excel.pack_forget()
            self.btn_export_excel.pack_forget()
            self.btn_preview_pdf.pack(padx=20, pady=4, fill="x", after=self.lbl_previews)
            self.btn_export_pdf.pack(padx=20, pady=4, fill="x", after=self.lbl_exports)
            self.btn_preview_pdf.configure(text="🔍 Preview Proforma (PDF)")
            self.btn_export_pdf.configure(text="📜 Save & Export Proforma PDF")

    def workspace_setup(self):
        self.tabs_container = ctk.CTkTabview(self.main_workspace, height=760)
        self.tabs_container.pack(fill="both", expand=True)

        tab_matrix = self.tabs_container.add("Matrix Selection")
        
        top_action_bar = ctk.CTkFrame(tab_matrix, fg_color="#eef2f7", height=50, corner_radius=8)
        top_action_bar.pack(fill="x", padx=5, pady=(5, 10))
        top_action_bar.pack_propagate(False)

        self.matrix_search_var = ctk.StringVar()
        self.matrix_search_var.trace_add("write", self.filter_matrix_by_product_name)
        
        ent_matrix_search = ctk.CTkEntry(
            top_action_bar, 
            textvariable=self.matrix_search_var, 
            placeholder_text="🔍 Search product name or code...", 
            width=320,
            height=35,
            corner_radius=10,
            border_width=2,
            border_color="#3b82f6",
            fg_color="#ffffff",
            placeholder_text_color="#94a3b8",
            text_color="#0f172a",
            font=("Segoe UI", 13)
        )
        ent_matrix_search.pack(side="left", padx=15, pady=7)

        btn_delete_item = ctk.CTkButton(top_action_bar, text="❌ Delete Selected", font=("Arial", 12, "bold"), fg_color="#cc3333", hover_color="#992222", width=130, command=self.delete_item_global_trigger)
        btn_delete_item.pack(side="right", padx=10, pady=10)
        
        btn_edit_item = ctk.CTkButton(top_action_bar, text="📝 Edit Selected", font=("Arial", 12, "bold"), fg_color="#1f538d", hover_color="#153a63", width=120, command=self.edit_item_global_trigger)
        btn_edit_item.pack(side="right", padx=10, pady=10)
        
        btn_add_item = ctk.CTkButton(top_action_bar, text="➕ Add New Item", font=("Arial", 12, "bold"), fg_color="#228B22", hover_color="#006400", width=120, command=self.add_item_global_trigger)
        btn_add_item.pack(side="right", padx=10, pady=10)

        # Refresh Workspace Button setup
        btn_refresh_system = ctk.CTkButton(
            top_action_bar, 
            text="🔄 Refresh System", 
            font=("Arial", 12, "bold"), 
            fg_color="#5a6268",        # Slate gray color takay positive actions se different lagay
            hover_color="#495057", 
            width=130, 
            command=self.refresh_entire_workspace
        )
        btn_refresh_system.pack(side="right", padx=10, pady=10)

        

        self.scroll_grid = ctk.CTkScrollableFrame(tab_matrix, fg_color="white", corner_radius=8, border_width=1, border_color="#dbdbdb")
        self.scroll_grid.pack(fill="both", expand=True, padx=5, pady=5)

        self.tabview_sub = ctk.CTkTabview(self.scroll_grid, height=590, command=self.reset_matrix_search_on_tab_change)
        self.tabview_sub.pack(fill="both", expand=True)

        for cat in self.inventory_db.keys():
            tab = self.tabview_sub.add(cat)
            
            frame_th = ctk.CTkFrame(tab, fg_color="transparent")
            frame_th.pack(fill="x", padx=10, pady=6)
            
            ctk.CTkLabel(frame_th, text="Sel", width=45, font=("Arial", 12, "bold"), anchor="center").pack(side="left")
            ctk.CTkLabel(frame_th, text="Items ", width=340, anchor="w", font=("Arial", 12, "bold")).pack(side="left", padx=10)
            ctk.CTkLabel(frame_th, text="Code Configuration", width=170, font=("Arial", 12, "bold"), anchor="w").pack(side="left", padx=5)
            ctk.CTkLabel(frame_th, text="Unit Price (€)", width=160, font=("Arial", 12, "bold"), anchor="w").pack(side="left", padx=5) 
            ctk.CTkLabel(frame_th, text="Qty", width=65, font=("Arial", 12, "bold"), anchor="center").pack(side="left", padx=5)
            
            sub_scroll = ctk.CTkScrollableFrame(tab, height=440)
            sub_scroll.pack(fill="both", expand=True)
            
            self.sub_scroll_frames[cat] = sub_scroll
            self.grid_inputs[cat] = []
            
            self.populate_category_rows(cat)

        self.tab_dashboard = self.tabs_container.add("📊 Dashboard & History Tracking")
        self.setup_dashboard_view()

    def populate_category_rows(self, cat):
        for widget in self.sub_scroll_frames[cat].winfo_children():
            widget.destroy()
        self.grid_inputs[cat].clear()

        products = self.inventory_db.get(cat, [])
        for prod in products:
            row = ctk.CTkFrame(self.sub_scroll_frames[cat], fg_color="transparent", height=45)
            row.pack(fill="x", padx=5, pady=4)
            row.pack_propagate(False)

            chk_frame = ctk.CTkFrame(row, fg_color="transparent", width=45, height=45)
            chk_frame.pack(side="left")
            chk_frame.pack_propagate(False)
            chk = ctk.CTkCheckBox(chk_frame, text="", width=24, height=24)
            chk.pack(expand=True)

            lbl_item = ctk.CTkLabel(row, text=prod["name"], width=340, height=45, anchor="w", font=("Arial", 13))
            lbl_item.pack(side="left", padx=10)

            code_compound_frame = ctk.CTkFrame(row, fg_color="transparent", width=170, height=45)
            code_compound_frame.pack(side="left", padx=5)
            code_compound_frame.pack_propagate(False)
            
            chk_code_show = ctk.CTkCheckBox(code_compound_frame, text="", width=20, height=20)
            chk_code_show.select()
            chk_code_show.pack(side="left", padx=(0, 5))

            ent_code = ctk.CTkEntry(code_compound_frame, width=115, height=30, font=("Arial", 12, "bold"), text_color="#1e293b", fg_color="#f1f5f9")
            ent_code.insert(0, prod["code"])
            # Default normal state taake functions data read/write aram se kar sakein
            ent_code.configure(state="normal") 
            ent_code.pack(side="left", expand=True, fill="x")

            price_compound_frame = ctk.CTkFrame(row, fg_color="transparent", width=160, height=45)
            price_compound_frame.pack(side="left", padx=5)
            price_compound_frame.pack_propagate(False)

            chk_price_show = ctk.CTkCheckBox(price_compound_frame, text="", width=20, height=20)
            chk_price_show.select() 
            chk_price_show.pack(side="left", padx=(0, 5))

            ent_price = ctk.CTkEntry(price_compound_frame, width=110, height=30, font=("Arial", 12, "bold"), justify="center", text_color="#1e293b", fg_color="#f1f5f9")
            ent_price.insert(0, prod.get("price", ""))
            ent_price.configure(state="normal")
            ent_price.pack(side="left", expand=True, fill="x")

            qty_frame = ctk.CTkFrame(row, fg_color="transparent", width=65, height=45)
            qty_frame.pack(side="left", padx=5)
            qty_frame.pack_propagate(False)
            ent_qty = ctk.CTkEntry(qty_frame, width=55, height=30, placeholder_text="1", font=("Arial", 12), justify="center")
            ent_qty.pack(expand=True)

            # --- DYNAMIC TOGGLE LOGIC FOR ENTRY BOXES ---
            # Jab Code Checkbox tick/untick ho toh UI entry box enable ya disable ho jaye
            def toggle_code_entry(e_code=ent_code, c_show=chk_code_show):
                if c_show.get() == 1:
                    e_code.configure(state="normal", fg_color="#f1f5f9", text_color="#1e293b")
                else:
                    e_code.configure(state="disabled", fg_color="#e2e8f0", text_color="#94a3b8")

            # Jab Price Checkbox tick/untick ho toh UI entry price box enable/disable ho aur Qty field behavior sync karein
            def toggle_price_entry(e_price=ent_price, e_qty=ent_qty, p_show=chk_price_show):
                if p_show.get() == 1:
                    e_price.configure(state="normal", fg_color="#f1f5f9", text_color="#1e293b")
                    e_qty.configure(state="normal", fg_color="#ffffff")
                else:
                    e_price.configure(state="disabled", fg_color="#e2e8f0", text_color="#94a3b8")
                    e_qty.configure(state="disabled", fg_color="#e2e8f0")

            chk_code_show.configure(command=toggle_code_entry)
            chk_price_show.configure(command=toggle_price_entry)

            self.grid_inputs[cat].append({
                "id": prod["id"], "name": prod["name"], "specs": prod["specs"], "code": prod["code"],
                "checkbox": chk, "code_entry": ent_code, "code_show": chk_code_show, 
                "qty": ent_qty, "price": ent_price, "price_show": chk_price_show,
                "row_frame": row  
            })

    def add_item_global_trigger(self):
        cat = self.tabview_sub.get()
        if cat: self.add_item_trigger(cat)

    def edit_item_global_trigger(self):
        cat = self.tabview_sub.get()
        if cat: self.edit_item_trigger(cat)

    def delete_item_global_trigger(self):
        cat = self.tabview_sub.get()
        if cat: self.delete_item_trigger(cat)

    def filter_matrix_by_product_name(self, *args):
        active_cat = self.tabview_sub.get()
        if not active_cat or active_cat not in self.grid_inputs:
            return
            
        search_query = self.matrix_search_var.get().strip().lower()
        search_tokens = search_query.split()
        
        for item in self.grid_inputs[active_cat]:
            product_name = item["name"].lower()
            product_code = item["code"].lower()
            
            matches_all_tokens = True
            for token in search_tokens:
                if token not in product_name and token not in product_code:
                    matches_all_tokens = False
                    break
            
            if matches_all_tokens:
                item["row_frame"].pack(fill="x", padx=5, pady=4)
            else:
                item["row_frame"].pack_forget()

    def reset_matrix_search_on_tab_change(self):
        self.matrix_search_var.set("")

    def add_item_trigger(self, cat):
            dialog = CustomInputDialog(self, f"Add Product to {cat}", [
                ("Item Nomenclature Name", "name"),
                ("Item Reference Code (e.g., BR-NEW)", "code"),
                ("Base Price (€)", "price"),
                ("Detailed Specification Description (PDF Only)", "specs")
            ])
            self.wait_window(dialog)
            
            # 1. Check karein agar user ne cancel ya cross (X) kiya ho
            if not dialog.result:
                return
                
            new_name = dialog.result.get("name", "").strip()
            new_code = dialog.result.get("code", "").strip()

            # 2. Validation: Name khali nahi hona chahiye
            if not new_name:
                messagebox.showerror("Validation Error", "Product Name dalna lazmi hai! Khali product add nahi ho sakta.")
                return

            # 3. STRICT DUPLICATION CHECK (Case-insensitive aur Whitespace-proof)
            if cat in self.inventory_db:
                for item in self.inventory_db[cat]:
                    existing_name = str(item.get("name", "")).strip().lower()
                    existing_code = str(item.get("code", "")).strip().lower()
                    
                    # Name duplicate check
                    if existing_name == new_name.lower():
                        messagebox.showerror("Duplicate Product", f"Problem! '{new_name}' Product already exist.")
                        return  # Code yahi ruk jayega, aage nahi barhega

                    # Code duplicate check (Agar code field khali na ho)
                    if new_code and existing_code == new_code.lower():
                        messagebox.showerror("Duplicate Code", f"Problem! '{new_code}' Product already exist.")
                        return  # Code yahi ruk jayega, aage nahi barhega

            # 4. AGAR APNE CHECKS PASS HO JAYEN TOU SAVE KAREIN
            new_item = {
                "id": f"gen.{datetime.now().timestamp()}",
                "name": new_name,
                "code": new_code,
                "price": dialog.result.get("price", "").strip(),
                "specs": dialog.result.get("specs", "").strip()
            }
            
            if cat not in self.inventory_db:
                self.inventory_db[cat] = []
                
            self.inventory_db[cat].append(new_item)
            
            # State aur GUI synchronize karein
            save_inventory_to_db(self.inventory_db)
            self.populate_category_rows(cat)
            
            messagebox.showinfo("Matrix Synchronized", f"'{new_name}' safely injected into {cat} database records.")

    def edit_item_trigger(self, cat):
        selected_index = None
        for idx, item_ui in enumerate(self.grid_inputs[cat]):
            if item_ui["checkbox"].get() == 1:
                if selected_index is not None:
                    messagebox.showwarning("Multiple Selection", "Please check/select only ONE item to edit at a time.")
                    return
                selected_index = idx
                
        if selected_index is None:
            messagebox.showwarning("Selection Empty", "Please check/select the checkbox of the single item you want to edit.")
            return

        current_item = self.inventory_db[cat][selected_index]
        initial_vals = {
            "name": current_item["name"],
            "code": current_item["code"],
            "price": current_item.get("price", ""),
            "specs": current_item["specs"]
        }

        dialog = CustomInputDialog(self, f"Modify Record inside {cat}", [
            ("Item Nomenclature Name", "name"),
            ("Item Reference Code", "code"),
            ("Base Price (€)", "price"),
            ("Detailed Specification Description", "specs")
        ], initial_values=initial_vals)
        self.wait_window(dialog)

        if dialog.result:
            self.inventory_db[cat][selected_index]["name"] = dialog.result["name"]
            self.inventory_db[cat][selected_index]["code"] = dialog.result["code"]
            self.inventory_db[cat][selected_index]["price"] = dialog.result["price"]
            self.inventory_db[cat][selected_index]["specs"] = dialog.result["specs"]
            
            save_inventory_to_db(self.inventory_db)
            self.populate_category_rows(cat)
            messagebox.showinfo("Success", "Core structural specification item records updated successfully.")

    def delete_item_trigger(self, cat):
        items_to_remove = []
        indices_to_remove = []
        
        for idx, item_ui in enumerate(self.grid_inputs[cat]):
            if item_ui["checkbox"].get() == 1:
                items_to_remove.append(item_ui["name"])
                indices_to_remove.append(idx)
                
        if not items_to_remove:
            messagebox.showwarning("Selection Empty", "Please check/select the checkbox of items you intend to erase.")
            return
            
        confirm = messagebox.askyesno("Confirm Purge", f"Are you sure you want to permanently erase the following {len(items_to_remove)} items?\n\n" + ", ".join(items_to_remove))
        if confirm:
            for idx in sorted(indices_to_remove, reverse=True):
                del self.inventory_db[cat][idx]
            save_inventory_to_db(self.inventory_db)
            self.populate_category_rows(cat)
            messagebox.showinfo("Purge Success", "Selected machine models deleted.")

    def setup_dashboard_view(self):
        frame_stats = ctk.CTkFrame(self.tab_dashboard, fg_color="transparent")
        frame_stats.pack(fill="x", padx=15, pady=10)
        
        self.card_quotations = ctk.CTkButton(frame_stats, text="Quotations Mapped\n0 Documents", font=("Arial", 14, "bold"), fg_color="#1f538d", height=70, state="disabled")
        self.card_quotations.pack(side="left", fill="x", expand=True, padx=5)

        self.card_proformas = ctk.CTkButton(frame_stats, text="Proformas Executed\n0 Documents", font=("Arial", 14, "bold"), fg_color="#228B22", height=70, state="disabled")
        self.card_proformas.pack(side="left", fill="x", expand=True, padx=5)

        log_header_panel = ctk.CTkFrame(self.tab_dashboard, fg_color="transparent", height=45)
        log_header_panel.pack(fill="x", padx=15, pady=(5, 0))

        self.db_search_var = ctk.StringVar()
        self.db_search_var.trace_add("write", self.filter_dashboard_logs_realtime)

        ent_db_search = ctk.CTkEntry(
            log_header_panel, 
            textvariable=self.db_search_var, 
            placeholder_text="🔍 Search ref, client or project...", 
            width=280,
            height=35,
            corner_radius=10,
            border_width=2,
            border_color="#228B22",
            fg_color="#ffffff",
            placeholder_text_color="#94a3b8",
            text_color="#0f172a",
            font=("Segoe UI", 12)
        )
        ent_db_search.pack(side="right", padx=10, pady=5)

        self.history_filters = ctk.CTkTabview(self.tab_dashboard, height=440)
        self.history_filters.pack(fill="both", expand=True, padx=15, pady=5)
        
        tab_q_log = self.history_filters.add("📋 Quotation Records Log")
        tab_p_log = self.history_filters.add("📜 Proforma Invoice Log")

        style = ttk.Style()
        style.theme_use("clam")
        style.configure("Treeview", background="white", fieldbackground="white", rowheight=32, font=("Arial", 12))
        style.configure("Treeview.Heading", background="#f2f5f9", font=("Arial", 12, "bold"), anchor="w")
        style.map("Treeview", background=[("selected", "#b3d1ff")], foreground=[("selected", "black")])

        self.tree_q = ttk.Treeview(tab_q_log, columns=("Ref", "Client", "Project", "Date", "Month", "Amount"), show="headings")
        self.setup_tree_columns(self.tree_q)
        self.tree_q.pack(fill="both", expand=True, padx=5, pady=5)

        self.tree_p = ttk.Treeview(tab_p_log, columns=("Ref", "Client", "Project", "Date", "Month", "Amount"), show="headings")
        self.setup_tree_columns(self.tree_p)
        self.tree_p.pack(fill="both", expand=True, padx=5, pady=5)

        frame_actions = ctk.CTkFrame(self.tab_dashboard, fg_color="transparent")
        frame_actions.pack(fill="x", padx=15, pady=10)
        
        ctk.CTkButton(frame_actions, text="🔄 Refresh Logs", font=("Arial", 12, "bold"), width=140, command=self.refresh_dashboard_data).pack(side="left", padx=5)
        ctk.CTkButton(frame_actions, text="📂 Load Selected Revision Record", font=("Arial", 13, "bold"), fg_color="#ffcc00", text_color="black", hover_color="#e6b800", command=self.load_selected_record).pack(side="left", padx=5)
        ctk.CTkButton(frame_actions, text="🗑️ Delete Selected Record", font=("Arial", 13, "bold"), fg_color="#cc3333", text_color="white", hover_color="#992222", command=self.delete_selected_record).pack(side="left", padx=5)

        self.cached_q_logs = []
        self.cached_p_logs = []
        self.refresh_dashboard_data()

    def setup_tree_columns(self, tree):
        tree.heading("Ref", text="  Reference / Revision Code", anchor="w")
        tree.heading("Client", text="  Client Factory Name", anchor="w")
        tree.heading("Project", text="  Project Module", anchor="w")
        tree.heading("Date", text="  Timestamp", anchor="w")
        tree.heading("Month", text="  Tracking Month", anchor="w")
        tree.heading("Amount", text="  Total Value", anchor="w")

        tree.column("Ref", width=180, anchor="w")
        tree.column("Client", width=260, anchor="w")
        tree.column("Project", width=160, anchor="w")
        tree.column("Date", width=120, anchor="w")
        tree.column("Month", width=130, anchor="w")
        tree.column("Amount", width=130, anchor="w")

    def trigger_system_logout(self):
        ans = messagebox.askyesno("Confirm Logout", "Are you sure you want to log out and lock the active workspace?")
        if ans:
            self.app_loaded = False 
            if self.sidebar:
                self.sidebar.grid_forget()
                self.sidebar.destroy()
            if self.main_workspace:
                self.main_workspace.grid_forget()
                self.main_workspace.destroy()
            self.sidebar = None
            self.main_workspace = None
            self.login_view.pack(expand=True, fill="both")
            self.login_view.clear_fields()
            self.login_view.resize_background()
            messagebox.showinfo("Session Terminated", "Workspace safely locked.")
    def get_parsed_data(self):
            grouped = {}
            has_prices = False
            grand_total = 0.0
            
            if not hasattr(self, 'grid_inputs') or not self.grid_inputs:
                return grouped, has_prices, grand_total
                
            for cat, items in self.grid_inputs.items():
                selected = []
                for item in items:
                    # 1. Checkbox validation for selection
                    main_checked = False
                    if "checkbox" in item:
                        main_checked = (item["checkbox"].get() == 1)
                    elif "sel" in item:
                        main_checked = (item["sel"].get() == 1)
                        
                    if main_checked:
                        # 2. Extract visibility flags dynamically (Fallback to True if missing)
                        is_price_visible = True
                        if "price_show" in item:
                            is_price_visible = (item["price_show"].get() == 1)
                        elif "price_visible" in item and hasattr(item["price_visible"], "get"):
                            is_price_visible = (item["price_visible"].get() == 1)
                            
                        is_code_visible = True
                        if "code_show" in item:
                            is_code_visible = (item["code_show"].get() == 1)
                        elif "code_visible" in item and hasattr(item["code_visible"], "get"):
                            is_code_visible = (item["code_visible"].get() == 1)
                        
                        # 3. Extract Code Value safely
                        code_val = ""
                        if "code_entry" in item: 
                            code_val = item["code_entry"].get().strip()
                        elif "code" in item and hasattr(item["code"], "get"): 
                            code_val = item["code"].get().strip()
                        if not is_code_visible:
                            code_val = "" # Agar visible nahi hai toh khali bhejein
                        
                        # 4. Extract Price String dynamically from UI inputs
                        price_str = ""
                        possible_price_keys = ["price_entry", "price", "unit_price", "price_field"]
                        for k in possible_price_keys:
                            if k in item and hasattr(item[k], "get"):
                                price_str = str(item[k].get()).strip()
                                break
                        
                        # Fallback: Agar screen par input khali hai toh database ka default check karein
                        if not price_str and "price_default" in item:
                            price_str = str(item["price_default"]).strip()
                        
                        # 5. Extract Qty Value safely (Hamesha number rkhna hai calculations k liye)
                        qty_val = 1
                        if "qty" in item:
                            try:
                                s_qty = str(item["qty"].get()).strip()
                                qty_val = int(s_qty) if s_qty else 1
                            except (ValueError, AttributeError):
                                qty_val = 1
                        
                        # 6. Process prices safely (Clean commas/currency symbols if any)
                        price = 0.0
                        if price_str:
                            try:
                                # Sanitize numeric string (comma saaf karein taake float converter error na de)
                                cleaned_price = price_str.replace(",", "").replace("€", "").strip()
                                price = float(cleaned_price) if cleaned_price else 0.0
                            except ValueError:
                                price = 0.0

                        if is_price_visible and price > 0:
                            has_prices = True
                            final_total = price * qty_val
                            grand_total += final_total
                        else:
                            # Agar price hide hai, toh UI visualization k liye data formatting empty rakhein
                            final_total = 0.0
                            if not is_price_visible:
                                price = ""
                                final_total = ""
                            
                        # 7. Package structured dictionary back to engine standard format
                        selected.append({
                            "id": item.get("id", ""), 
                            "name": item.get("name", ""), 
                            "code": code_val, 
                            "code_visible": is_code_visible,
                            "specs": item.get("specs", ""), 
                            "qty": qty_val if is_price_visible else "", 
                            "price": price, 
                            "total": final_total,
                            "price_visible": is_price_visible
                        })
                        
                if selected: 
                    grouped[cat] = selected
                    
            return grouped, has_prices, grand_total

    def auto_log_state_to_history(self, grand_total):
        ref_num = self.ent_offer.get().strip() or "UNTITLED_REF"
        doc_type = self.doc_type_var.get()
        current_month_folder = datetime.now().strftime("%B_%Y")
        sub_folder_name = "Quotations" if doc_type == "Quotation" else "Proformas"
        target_directory_path = os.path.join(BASE_HISTORY_DIR, current_month_folder, sub_folder_name)
        if not os.path.exists(target_directory_path):
            os.makedirs(target_directory_path)

        state_payload = {
            "type": doc_type, "factory": self.ent_factory.get().strip(),
            "address": self.ent_address.get().strip() if doc_type == "Proforma" else "",
            "offer": ref_num, "project": self.ent_project.get().strip() if doc_type == "Proforma" else "",
            "date": datetime.now().strftime("%d.%m.%Y"), "month": current_month_folder,
            "total_value": grand_total, "selections": {}
        }
        for cat, items in self.grid_inputs.items():
            state_payload["selections"][cat] = [{
                "checked": x["checkbox"].get(), 
                "code": x["code_entry"].get(), 
                "code_show": x["code_show"].get(),
                "qty": x["qty"].get(), 
            
                "price": x["price"].get(),
                "price_show": x["price_show"].get()
            } for x in items]

        filename_prefix = f"{doc_type}_{ref_num}"
        existing_file = os.path.join(target_directory_path, f"{filename_prefix}.json")
        with open(existing_file, "w") as f: json.dump(state_payload, f, indent=4)
        self.refresh_dashboard_data()
        return existing_file

    def refresh_dashboard_data(self):
        for row in self.tree_q.get_children(): self.tree_q.delete(row)
        for row in self.tree_p.get_children(): self.tree_p.delete(row)
        
        self.cached_q_logs.clear()
        self.cached_p_logs.clear()
        
        q_count, p_count = 0, 0
        if os.path.exists(BASE_HISTORY_DIR):
            for month_dir in os.listdir(BASE_HISTORY_DIR):
                full_m_path = os.path.join(BASE_HISTORY_DIR, month_dir)
                if os.path.isdir(full_m_path):
                    q_folder = os.path.join(full_m_path, "Quotations")
                    if os.path.exists(q_folder):
                        for file in os.listdir(q_folder):
                            if file.endswith(".json"):
                                try:
                                    with open(os.path.join(q_folder, file), "r") as f: d = json.load(f)
                                    tokens = f"{d.get('offer','')} {d.get('factory','')} N/A (Quotation) {d.get('date','')} {d.get('month','')}".lower()
                                    vals = (d.get("offer"), d.get("factory"), "N/A (Quotation)", d.get("date"), d.get("month"), f"€ {d.get('total_value',0):,.2f}")
                                    self.cached_q_logs.append((vals, tokens))
                                    q_count += 1
                                except: pass
                    p_folder = os.path.join(full_m_path, "Proformas")
                    if os.path.exists(p_folder):
                        for file in os.listdir(p_folder):
                            if file.endswith(".json"):
                                try:
                                    with open(os.path.join(p_folder, file), "r") as f: d = json.load(f)
                                    tokens = f"{d.get('offer','')} {d.get('factory','')} {d.get('project','N/A')} {d.get('date','')} {d.get('month','')}".lower()
                                    vals = (d.get("offer"), d.get("factory"), d.get("project", "N/A"), d.get("date"), d.get("month"), f"€ {d.get('total_value',0):,.2f}")
                                    self.cached_p_logs.append((vals, tokens))
                                    p_count += 1
                                except: pass

        self.apply_dashboard_tree_renders(self.cached_q_logs, self.cached_p_logs)
        self.card_quotations.configure(text=f"Quotations Mapped\n{q_count} Documents")
        self.card_proformas.configure(text=f"Proformas Executed\n{p_count} Documents")

    def apply_dashboard_tree_renders(self, q_list, p_list):
        for row_data, _ in q_list: self.tree_q.insert("", "end", values=row_data)
        for row_data, _ in p_list: self.tree_p.insert("", "end", values=row_data)

    def filter_dashboard_logs_realtime(self, *args):
        query = self.db_search_var.get().strip().lower()
        query_tokens = query.split()

        for item in self.tree_q.get_children(): self.tree_q.delete(item)
        for item in self.tree_p.get_children(): self.tree_p.delete(item)

        filtered_q = [r for r in self.cached_q_logs if all(t in r[1] for t in query_tokens)]
        filtered_p = [r for r in self.cached_p_logs if all(t in r[1] for t in query_tokens)]
        self.apply_dashboard_tree_renders(filtered_q, filtered_p)

    def load_selected_record(self):
        active_tab = self.history_filters.get()
        if "Quotation" in active_tab:
            sel = self.tree_q.selection(); target_tree = self.tree_q; sub_dir_name = "Quotations"; doc_prefix = "Quotation"
        else:
            sel = self.tree_p.selection(); target_tree = self.tree_p; sub_dir_name = "Proformas"; doc_prefix = "Proforma"
        if not sel:
            messagebox.showwarning("Selection Missing", "Please select a specific history record line profile to load.")
            return
        item_vals = target_tree.item(sel[0], "values")
        target_ref = item_vals[0]; target_month = item_vals[4]
        target_path = os.path.join(BASE_HISTORY_DIR, target_month, sub_dir_name, f"{doc_prefix}_{target_ref}.json")
        if os.path.exists(target_path):
            with open(target_path, "r") as f: data = json.load(f)
            self.doc_type_var.set(data.get("type", "Quotation"))
            self.toggle_dynamic_fields()
            self.ent_factory.delete(0, "end"); self.ent_factory.insert(0, data.get("factory", ""))
            self.ent_offer.delete(0, "end"); self.ent_offer.insert(0, data.get("offer", ""))
            if data.get("type") == "Proforma":
                self.ent_address.delete(0, "end"); self.ent_address.insert(0, data.get("address", ""))
                self.ent_project.delete(0, "end"); self.ent_project.insert(0, data.get("project", ""))
            for cat, items in self.grid_inputs.items():
                saved = data.get("selections", {}).get(cat, [])
                for s, sv in zip(items, saved):
                    s["checkbox"].deselect()
                    if sv["checked"] == 1: s["checkbox"].select()
                    s["code_entry"].configure(state="normal"); s["code_entry"].delete(0, "end"); s["code_entry"].insert(0, sv.get("code", "")); s["code_entry"].configure(state="disabled")
                    s["code_show"].deselect(); 
                    if sv.get("code_show", 1) == 1: s["code_show"].select()
                    
                    s["price_show"].deselect()
                    if sv.get("price_show", 1) == 1: s["price_show"].select()
                    
                    s["price"].configure(state="normal")
                    s["price"].delete(0, "end")
                    s["price"].insert(0, sv.get("price", ""))
                    s["price"].configure(state="disabled")
                    
                    s["qty"].delete(0, "end"); s["qty"].insert(0, sv["qty"])
                    
            self.tabs_container.set("Matrix Selection"); messagebox.showinfo("State Restored", f"Profile Entry '{target_ref}' loaded.")

    def delete_selected_record(self):
        active_tab = self.history_filters.get()
        if "Quotation" in active_tab:
            sel = self.tree_q.selection(); target_tree = self.tree_q; sub_dir_name = "Quotations"; doc_prefix = "Quotation"
        else:
            sel = self.tree_p.selection(); target_tree = self.tree_p; sub_dir_name = "Proformas"; doc_prefix = "Proforma"
        if not sel:
            messagebox.showwarning("Selection Missing", "Please select a specific history record line profile to delete.")
            return
        item_vals = target_tree.item(sel[0], "values")
        target_ref = item_vals[0]; target_month = item_vals[4]
        target_path = os.path.join(BASE_HISTORY_DIR, target_month, sub_dir_name, f"{doc_prefix}_{target_ref}.json")
        if os.path.exists(target_path): os.remove(target_path); messagebox.showinfo("Deleted", "Record profile successfully removed."); self.refresh_dashboard_data()

    def trigger_pdf_preview(self):
        try:
            grouped, _, _ = self.get_parsed_data() 
            if not grouped: 
                messagebox.showwarning("Selection Missing", "No items selected.")
                return
            temp_preview_path = os.path.join(os.environ.get('TEMP', ''), "trutzschler_temp_preview.pdf")
            self.build_pdf_document_engine(temp_preview_path)
            webbrowser.open(temp_preview_path)
        except Exception as e:
            messagebox.showerror("Preview Error", f"PDF Preview generate karne mein masala aya:\n{str(e)}")

    def trigger_excel_preview(self):
        try:
            grouped, _, _ = self.get_parsed_data() 
            if not grouped: 
                messagebox.showwarning("Selection Missing", "No items selected.")
                return
            temp_xl_path = os.path.join(os.environ.get('TEMP', ''), "trutzschler_temp_preview.xlsx")
            self.build_excel_sheet_engine(temp_xl_path)
            webbrowser.open(temp_xl_path)
        except Exception as e:
            messagebox.showerror("Preview Error", f"Excel Preview generate karne mein masala aya:\n{str(e)}")

    def generate_pdf_proforma(self):
        try:
            grouped, _, grand_total = self.get_parsed_data() 
            if not grouped: 
                messagebox.showwarning("Empty Selection", "No components selected.")
                return
            save_path = filedialog.asksaveasfilename(defaultextension=".pdf", filetypes=[("PDF Document Replica", "*.pdf")])
            if not save_path: 
                return
            self.build_pdf_document_engine(save_path)
            self.auto_log_state_to_history(grand_total)
            messagebox.showinfo("Export Success", "Official layout exported!")
        except Exception as e:
            messagebox.showerror("Export Error", f"PDF Export fail ho gya:\n{str(e)}")

    def generate_excel_format(self):
        try:
            grouped, _, grand_total = self.get_parsed_data() 
            if not grouped: 
                messagebox.showwarning("Empty Structure", "No machinery configurations mapped.")
                return
            save_p = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Sheet", "*.xlsx")])
            if not save_p: 
                return
            self.build_excel_sheet_engine(save_p)
            self.auto_log_state_to_history(grand_total)
            messagebox.showinfo("Export Success", "Spreadsheet generated!")
        except Exception as e:
            messagebox.showerror("Export Error", f"Excel Export fail ho gya:\n{str(e)}")

    def refresh_entire_workspace(self):
            # 1. Left panel ki document configuration fields ko clear karein
            if hasattr(self, 'ent_factory') and self.ent_factory: 
                self.ent_factory.delete(0, 'end')
            if hasattr(self, 'ent_address') and self.ent_address: 
                self.ent_address.delete(0, 'end')
            if hasattr(self, 'ent_offer') and self.ent_offer: 
                self.ent_offer.delete(0, 'end')
            if hasattr(self, 'ent_project') and self.ent_project: 
                self.ent_project.delete(0, 'end')
            
            # Doc type dropdown ko default (Quotation) ya Proforma par reset karein
            if hasattr(self, 'doc_type_var') and self.doc_type_var: 
                self.doc_type_var.set("Quotation")
                self.toggle_dynamic_fields() # Left panel ke visibility fields ko sync karne ke liye
                
            # Search bar ko bhi khali kar dein agar usme kuch likha ho
            if hasattr(self, 'matrix_search_var') and self.matrix_search_var:
                self.matrix_search_var.set("")

            # 2. Saari categories ke items ko database se bilkul fresh state me repopulate (reset) karein
            if hasattr(self, 'inventory_db') and self.inventory_db:
                for cat in self.inventory_db.keys():
                    try:
                        self.populate_category_rows(cat)
                    except Exception as e:
                        print(f"Error resetting category {cat}: {e}")
                        
            # Success message box taake pata chale system refresh ho gya h
            messagebox.showinfo("System Refreshed", " successfully syestem refreshed.")


    # --- FIRST PAGE FULLY CUSTOMIZED PROFORMA INVOICE RE-ENGINEERING ---
    # --- FIRST PAGE FULLY CUSTOMIZED PROFORMA INVOICE RE-ENGINEERING ---
    def build_pdf_document_engine(self, target_path):
        factory = self.ent_factory.get().strip() or "MAHMOOD TEXTILE"
        doc_type = self.doc_type_var.get()
        
        if doc_type == "Quotation": 
            address = ""
            proj_name = ""
        else:
            address = self.ent_address.get().strip() or "SURVEY # 31,32,33,34, KARACHI, PAKISTAN"
            proj_name = self.ent_project.get().strip() or "Draw Frames"
            
        offer_num = self.ent_offer.get().strip() or "02-013106-00"
        grouped, has_prices, grand_total = self.get_parsed_data()
        
        doc = SimpleDocTemplate(target_path, pagesize=letter, leftMargin=54, rightMargin=54, topMargin=90, bottomMargin=110)
        
        s_normal = ParagraphStyle('Norm', fontName='Helvetica', fontSize=9, leading=14, textColor=colors.HexColor("#222222"))
        s_bold = ParagraphStyle('Bld', fontName='Helvetica-Bold', fontSize=9, leading=14)
        s_factory = ParagraphStyle('Fac', fontName='Helvetica-Bold', fontSize=12, leading=16, textColor=colors.HexColor("#111111"))
        s_title = ParagraphStyle('Ttl', fontName='Helvetica-Bold', fontSize=22, leading=26, spaceBefore=15, spaceAfter=5)
        s_sec = ParagraphStyle('Sec', fontName='Helvetica-Bold', fontSize=12, leading=16, spaceBefore=15, spaceAfter=10, textColor=colors.HexColor("#003399"))
        
        story = [Spacer(1, 15)]
        
        left_block = [Paragraph(factory.upper(), s_factory)]
        if address: 
            left_block.append(Paragraph(address, s_normal))
        else: 
            left_block.append(Spacer(1, 1))
            
        right_block = [
            Paragraph("<b>Official in charge:</b> Gerold Gonska", s_normal),
            Paragraph("<b>Department:</b> Sales", s_normal),
            Paragraph("<b>Fon:</b> +49 2166 607-213", s_normal),
            Paragraph("<b>Fax:</b> +49 2166 607-799", s_normal),
            Paragraph("<b>E-mail:</b> gerold.gonska@truetzschler.de", s_normal),
            Paragraph("<b>Website:</b> www.truetzschler.com", s_normal)
        ]
        
        header_table = Table([[left_block, right_block]], colWidths=[270, 230])
        header_table.setStyle(TableStyle([('VALIGN', (0,0), (-1,-1), 'TOP'), ('PADDING', (0,0), (-1,-1), 0)]))
        story.append(header_table)
        story.append(Spacer(1, 25))
        
        story.append(Paragraph(doc_type, s_title))
        story.append(Paragraph(f"<b>No.</b> {offer_num}", ParagraphStyle('OffNum', fontName='Helvetica', fontSize=12, leading=16, textColor=colors.HexColor("#444444"))))
        story.append(Spacer(1, 15))
        
        meta_left = []
        if proj_name:
            meta_left.append(Paragraph("<b>for your Project:</b>", s_normal))
            meta_left.append(Paragraph(proj_name, s_bold))
        else: 
            meta_left.append(Spacer(1, 1))
            
        meta_meta = [
            [meta_left, Paragraph(f"<b>Date:</b> {datetime.now().strftime('%d.%m.%Y')}", s_normal)],
            [Spacer(1, 12), ""],
            [Paragraph("<b>Representative:</b><br/>MACHPART INTERNATIONAL<br/>43-G, Gulberg-II<br/>Lahore<br/>Pakistan", s_normal), ""]
        ]
        
        t_meta = Table(meta_meta, colWidths=[270, 230])
        t_meta.setStyle(TableStyle([('VALIGN', (0,0), (-1,-1), 'TOP'), ('PADDING', (0,0), (-1,-1), 1)]))
        story.append(t_meta)
        
        if has_prices:
            story.append(PageBreak())
            story.append(Paragraph("Commercial Price Summary Breakdown", s_sec))
            sum_data = [["Section Description Details Nomenclature", "Net Price (EUR)"]]
            for cat, items in grouped.items():
                cat_subtotal = sum(x["total"] for x in items if isinstance(x["total"], float))
                if cat_subtotal > 0:
                    sum_data.append([f"Subtotal System Machinery Section — {cat}", f"€ {cat_subtotal:,.2f}"])
            sum_data.append(["TOTAL OVERALL SUMMARY VALUE OFFER", f"€ {grand_total:,.2f}"])
            t_sum = Table(sum_data, colWidths=[370, 130])
            t_sum.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#f8f9fa")), ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor("#cccccc")),
                ('ALIGN', (1,0), (1,-1), 'RIGHT'), ('FONTNAME', (0,-1), (-1,-1), 'Helvetica-Bold'), ('PADDING', (0,0), (-1,-1), 8)
            ]))
            story.append(t_sum)
            
        for cat, items in grouped.items():
            story.append(PageBreak())
            story.append(Paragraph(f"Technical Specification Module — {cat}", s_sec))
            
            cat_has_prices = any(x["price_visible"] and isinstance(x["price"], float) for x in items)
            any_code_visible = any(it["code_visible"] for it in items)
            
            if cat_has_prices:
                if any_code_visible:
                    headers = ["Code", "Subassembly Specification Description Details", "Qty", "Price", "Total"]
                    widths = [65, 230, 25, 90, 90]
                else:
                    headers = ["Subassembly Specification Description Details", "Qty", "Price", "Total"]
                    widths = [295, 25, 90, 90]
                
                table_rows = [headers]
                for it in items:
                    desc_p = Paragraph(f"<b>{it['name']}</b><br/><font color='#555555'>{it['specs']}</font>", s_normal)
                    q_text = str(it["qty"]) if it["qty"] != "" else ""
                    
                    p_text = f"€ {it['price']:,.2f}" if (it.get("price_visible") and isinstance(it.get("price"), float)) else ""
                    t_text = f"€ {it['total']:,.2f}" if (it.get("price_visible") and isinstance(it.get("total"), float)) else ""
                    
                    if any_code_visible:
                        table_rows.append([str(it["code"]), desc_p, q_text, p_text, t_text])
                    else:
                        table_rows.append([desc_p, q_text, p_text, t_text])
                        
                t_spec = Table(table_rows, colWidths=widths)
                t_spec.setStyle(TableStyle([
                    ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor("#dddddd")), ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#f2f5f9")),
                    ('VALIGN', (0,0), (-1,-1), 'TOP'), ('PADDING', (0,0), (-1,-1), 7), 
                    ('ALIGN', (-4,1), (-1,-1), 'CENTER')
                ]))
            else:
                if any_code_visible:
                    headers = ["Code", "Subassembly Specification Description Details"]
                    widths = [80, 420]
                else:
                    headers = ["Subassembly Specification Description Details"]
                    widths = [500]
                    
                table_rows = [headers]
                for it in items:
                    desc_p = Paragraph(f"<b>{it['name']}</b><br/><font color='#555555'>{it['specs']}</font>", s_normal)
                    if any_code_visible:
                        table_rows.append([str(it["code"]), desc_p])
                    else:
                        table_rows.append([desc_p])
                        
                t_spec = Table(table_rows, colWidths=widths)
                t_spec.setStyle(TableStyle([
                    ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor("#dddddd")), ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#f2f5f9")),
                    ('VALIGN', (0,0), (-1,-1), 'TOP'), ('PADDING', (0,0), (-1,-1), 8)
                ]))
                
            story.append(t_spec)
            
        doc.build(story, canvasmaker=NumberedCanvas)

          # --- FIRST PAGE FULLY DYNAMIC EXCEL ENGINE ---
    def build_excel_sheet_engine(self, target_path):
        factory = self.ent_factory.get().strip() or "MAHMOOD TEXTILE"
        offer_num = self.ent_offer.get().strip()
        doc_type = self.doc_type_var.get()
        
        # EXACT DATA LAYER THAT MAKES PDF PERFECT
        grouped, has_prices, grand_total = self.get_parsed_data()
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Proposal"
        ws.views.sheetView[0].showGridLines = True
        
        f_head = XLFont(name="Arial", size=15, bold=True)
        f_bold = XLFont(name="Arial", size=10, bold=True)
        f_normal = XLFont(name="Arial", size=10)
        fill_y = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        border_thin = Border(left=Side(style='thin', color='000000'), right=Side(style='thin', color='000000'), top=Side(style='thin', color='000000'), bottom=Side(style='thin', color='000000'))
        
        ws.merge_cells("A1:G1")
        ws["A1"] = f"{factory.upper()} - {doc_type.upper()}"
        ws["A1"].font = f_head
        ws["A1"].fill = fill_y
        ws["A1"].alignment = Alignment(horizontal="center")
        
        ws["A3"] = f"{doc_type} Pricing Framework Working"
        ws["D3"] = f"Ref No {offer_num}" if offer_num else ""
        ws["F2"] = "Based On New Price Matrix"
        ws["F2"].fill = fill_y
        ws["F3"] = datetime.now().strftime("%m/%d/%Y")
        
        current_row = 5
        global_any_code_visible = False

        for cat, items in grouped.items():
            if not items:
                continue
                
            # Category level mapping exactly matching PDF workflow
            cat_has_prices = any(x["price_visible"] and isinstance(x["price"], float) for x in items)
            any_code_visible = any(it["code_visible"] for it in items)
            
            if any_code_visible:
                global_any_code_visible = True
            
            if cat_has_prices:
                # --- MODE 1: PRICING PRESENT LAYOUT ---
                ws.cell(row=current_row, column=2, value="Quantity").font = f_bold
                c_sec = ws.cell(row=current_row, column=3, value=f"Extension of the {cat} Installation")
                c_sec.font = f_bold
                c_sec.fill = fill_y
                
                if any_code_visible:
                    ws.cell(row=current_row, column=4, value="CODE").font = f_bold
                    ws.cell(row=current_row, column=5, value="Unit Price").font = f_bold
                    ws.cell(row=current_row, column=6, value="Total Price").font = f_bold
                    max_col = 7
                else:
                    ws.cell(row=current_row, column=4, value="Unit Price").font = f_bold
                    ws.cell(row=current_row, column=5, value="Total Price").font = f_bold
                    max_col = 6
                        
                for c in range(2, max_col): 
                    ws.cell(row=current_row, column=c).border = border_thin
                current_row += 1
                
                cat_total = 0
                for item in items:
                    ws.cell(row=current_row, column=2, value=item["qty"])
                    ws.cell(row=current_row, column=3, value=item["name"])
                    
                    if any_code_visible:
                        ws.cell(row=current_row, column=4, value=str(item["code"]) if item.get("code_visible") else None)
                        p_col, t_col = 5, 6
                    else:
                        p_col, t_col = 4, 5
                        
                    if item.get("price_visible") and isinstance(item.get("price"), float):
                        ws.cell(row=current_row, column=p_col, value=item["price"]).number_format = '#,##0.00'
                        ws.cell(row=current_row, column=t_col, value=item["total"]).number_format = '#,##0.00'
                        cat_total += item["total"]
                    else:
                        ws.cell(row=current_row, column=p_col, value=None)
                        ws.cell(row=current_row, column=t_col, value=None)
                        
                    for c in range(2, max_col): 
                        ws.cell(row=current_row, column=c).font = f_normal
                        ws.cell(row=current_row, column=c).border = border_thin
                    current_row += 1
                    
                ws.cell(row=current_row, column=3, value=f"Line total {c_sec.value}").font = f_bold
                ws.cell(row=current_row, column=t_col, value=cat_total).font = f_bold
                ws.cell(row=current_row, column=t_col).number_format = '#,##0.00'
            else:
                # --- MODE 2: PURE SPECIFICATIONS (PRICES UNTICKED) ---
                c_sec = ws.cell(row=current_row, column=3, value=f"{cat} Specification Details")
                c_sec.font = f_bold
                c_sec.fill = fill_y
                
                # Header Border Fix (range ko 3 se 5 tak chalana hai agar code visible ho)
                end_border_col = 5 if any_code_visible else 4
                for c in range(3, end_border_col):
                    ws.cell(row=current_row, column=c).border = border_thin
                
                if any_code_visible:
                    ws.cell(row=current_row, column=4, value="CODE").font = f_bold
                    ws.cell(row=current_row, column=4).border = border_thin
                    
                current_row += 1
                for item in items:
                    ws.cell(row=current_row, column=3, value=item["name"]).font = f_normal
                    ws.cell(row=current_row, column=3).border = border_thin
                    
                    if any_code_visible:
                        ws.cell(row=current_row, column=4, value=str(item["code"]) if item.get("code_visible") else None).font = f_normal
                        ws.cell(row=current_row, column=4).border = border_thin
                    current_row += 1
                    
                ws.cell(row=current_row, column=3, value=f"{cat} Specification Mapped Successfully").font = f_bold
            
            current_row += 2
            
        # GRAND TOTAL: Strict numerical and boolean filter matching PDF structure
        if has_prices and isinstance(grand_total, (int, float)) and grand_total > 0: 
            final_col_idx = 6 if global_any_code_visible else 5
            ws.cell(row=current_row, column=3, value="GRAND TOTAL").font = f_bold
            ws.cell(row=current_row, column=final_col_idx, value=grand_total).font = f_bold
            ws.cell(row=current_row, column=final_col_idx).number_format = '#,##0.00'

        
            
        
            
       
            
        # --- COLUMNS AUTO-FIT & WIDTH SETTING ENGINE ---
            
            ws.column_dimensions['C'].width = 42  # Description / Item Name
            ws.column_dimensions['D'].width = 15  # Code column
            ws.column_dimensions['E'].width = 16  # Qty column
            ws.column_dimensions['F'].width = 18  # Unit Price column
            ws.column_dimensions['G'].width = 19  # Total Price / Grand Total column
            
            wb.save(target_path)

if __name__ == "__main__":
    app = MasterSystemApp()
    app.mainloop()